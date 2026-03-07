"""
Import / refresh research_static_reference from CSV/XLSX/PDF.

Usage:
  .venv/bin/python scripts/import_research_static_reference.py --source data/research_reference.csv --dry-run
  .venv/bin/python scripts/import_research_static_reference.py --source data/research_reference.xlsx
  .venv/bin/python scripts/import_research_static_reference.py --source data/research_reference.pdf
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import uuid
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from app.core.config import settings
from app.core.database import AsyncSessionLocal
from app.core.database import ResearchStaticReferenceModel
from sqlalchemy import and_, select, text


ALLOWED_DOMAINS = {
    "market_size",
    "industry_trends",
    "competitor_info",
    "policy_support",
}

PDF_HINT_BY_FILENAME = {
    "market": "market_size",
    "trend": "industry_trends",
    "동향": "industry_trends",
    "competitor": "competitor_info",
    "경쟁": "competitor_info",
    "policy": "policy_support",
    "지원": "policy_support",
}

DOMAIN_HINTS = {
    "market_size": ["시장규모", "market size", "시장 규모", "market_size", "매출"],
    "industry_trends": ["동향", "trend", "산업동향", "시장동향", "trend_trends"],
    "competitor_info": ["경쟁사", "경쟁", "competitor", "경쟁사 비교"],
    "policy_support": ["정책", "지원", "policy", "지원사업", "인증"],
}


def _normalize_bool(value: Optional[str]) -> bool:
    if value is None:
        return True
    if isinstance(value, bool):
        return value
    v = str(value).strip().lower()
    return v in {"1", "true", "y", "yes", "on", "활성", "v"}


def _normalize_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    v = str(value).strip()
    return v if v else None


def _parse_payload(value: Optional[str]) -> Dict[str, Any]:
    if value is None:
        return {}
    v = str(value).strip()
    if not v:
        return {}
    try:
        parsed = json.loads(v)
    except Exception:
        return {"raw": v}
    if isinstance(parsed, dict):
        return parsed
    return {"value": parsed}


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except Exception:
        return default


def _normalize_url(value: Optional[str]) -> Optional[str]:
    v = _normalize_text(value)
    if not v:
        return None
    if v.startswith("http://") or v.startswith("https://"):
        return v
    return None


def _make_row_id(row: Dict[str, Any]) -> str:
    key = "|".join(
        [
            str(row.get("domain") or ""),
            str(row.get("industry_code") or ""),
            str(row.get("tag") or ""),
            str(row.get("title") or ""),
        ]
    )
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, key))


def _coerce_row(
    row: Dict[str, Any],
    require_review_pass: bool,
    min_confidence: int,
) -> Tuple[Dict[str, Any], List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []

    domain = _normalize_text(row.get("domain") or row.get("도메인") or row.get("type"))
    if not domain:
        errors.append("domain is required")
    elif domain not in ALLOWED_DOMAINS:
        errors.append(f"invalid domain={domain}")

    title = _normalize_text(row.get("title") or row.get("제목") or row.get("name"))
    if not title:
        errors.append("title is required")

    if errors:
        return {}, errors

    industry_code = _normalize_text(row.get("industry_code") or row.get("industry") or row.get("industrycode"))
    tag = _normalize_text(row.get("tag") or row.get("태그") or row.get("category"))
    source_url = _normalize_url(row.get("source_url") or row.get("url") or row.get("source"))
    source_text = _normalize_text(
        row.get("source_text")
        or row.get("content")
        or row.get("summary")
        or row.get("description")
        or row.get("text")
        or ""
    )
    if source_url is None and _normalize_text(row.get("source_url") or row.get("url") or row.get("source")):
        warnings.append("invalid_url")

    parsed_confidence = _to_int(
        row.get("parsed_confidence") or row.get("confidence") or row.get("score"),
        default=0,
    )
    need_review = str(
        row.get("need_review") or row.get("review_needed") or ""
    ).strip().lower() in {"1", "true", "y", "yes", "on", "필수"}

    if parsed_confidence > 0 and parsed_confidence < min_confidence:
        warnings.append(f"low_confidence:{parsed_confidence}")
    if parsed_confidence <= 0 and min_confidence > 0:
        warnings.append(f"missing_confidence")

    if require_review_pass and need_review:
        warnings.append("needs_review")

    payload_json = _parse_payload(row.get("payload_json") or row.get("payload") or row.get("meta"))
    if not isinstance(payload_json, dict):
        payload_json = {"value": payload_json}
    payload_json = {
        **payload_json,
        "source_file": _normalize_text(row.get("source_file") or row.get("source")),
        "page": _normalize_text(row.get("page")),
        "parsed_confidence": parsed_confidence,
        "need_review": bool(need_review),
    }
    is_active = _normalize_bool(row.get("is_active") if "is_active" in row else row.get("active"))

    row_id = _normalize_text(row.get("id")) or _make_row_id(
        {
            "domain": domain,
            "industry_code": industry_code,
            "tag": tag,
            "title": title,
        }
    )

    return {
        "id": row_id,
        "domain": domain,
        "industry_code": industry_code,
        "tag": tag,
        "title": title,
        "source_url": source_url,
        "source_text": source_text,
        "payload_json": payload_json,
        "is_active": is_active,
    }, errors, warnings


def _is_blocked_quality_issue(warnings: List[str]) -> bool:
    for w in warnings:
        if w.startswith("needs_review") or w.startswith("low_confidence") or w.startswith("missing_confidence"):
            return True
    return False


def _detect_domain(text: str, filename: str) -> str:
    lowered = (text or "").lower()
    filename_lower = filename.lower()
    for keyword, domain in PDF_HINT_BY_FILENAME.items():
        if keyword in filename_lower:
            return domain

    best = ("market_size", 0)
    for domain, hints in DOMAIN_HINTS.items():
        score = 0
        for hint in hints:
            if hint.lower() in lowered:
                score += 1
        if score > best[1]:
            best = (domain, score)
    return best[0]


def _parse_pdf_labeled_line(line: str) -> Dict[str, str]:
    if ":" not in line:
        return {}
    key, value = line.split(":", 1)
    key = key.strip().lower()
    value = value.strip()
    if not value:
        return {}

    key_map = {
        "도메인": "domain",
        "domain": "domain",
        "산업코드": "industry_code",
        "industry_code": "industry_code",
        "industry": "industry_code",
        "tag": "tag",
        "카테고리": "tag",
        "유형": "tag",
        "title": "title",
        "제목": "title",
        "주제": "title",
        "출처": "source_url",
        "source": "source_url",
        "source_url": "source_url",
        "요약": "source_text",
        "summary": "source_text",
        "내용": "source_text",
    }
    mapped = key_map.get(key)
    if not mapped:
        return {}
    return {mapped: value}


def _split_pdf_lines(raw_text: str) -> List[str]:
    out: List[str] = []
    for row in raw_text.splitlines():
        row = row.strip()
        if not row:
            continue
        out.append(row)
    return out


def read_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    return rows


def read_xlsx(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl not installed. pip install openpyxl") from exc

    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or not rows[0]:
        raise ValueError("empty xlsx")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for values in rows[1:]:
        if not any(v is not None and str(v).strip() for v in values):
            continue
        item: Dict[str, Any] = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            item[h] = values[idx] if idx < len(values) else None
        out.append(item)
    return out


def read_pdf(path: Path) -> List[Dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "PDF 파서를 찾을 수 없습니다. requirements.txt에 pypdf 추가 후 pip install 필요."
        ) from exc

    try:
        reader = PdfReader(str(path))
    except Exception as exc:
        raise RuntimeError(f"PDF 열기 실패: {path}") from exc

    rows: List[Dict[str, Any]] = []
    for page_no, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if not text:
            continue

        lines = _split_pdf_lines(text)
        cur = {
            "title": f"{path.stem} p{page_no}",
            "tag": "pdf_import",
            "source_text": text[:4000],
            "payload_json": json.dumps(
                {"source_file": path.name, "page": page_no, "source_type": "pdf"}
            ),
            "source_url": None,
            "is_active": True,
        }

        for line in lines:
            mapped = _parse_pdf_labeled_line(line)
            for key, value in mapped.items():
                if key == "source_text" and not value:
                    continue
                cur[key] = value

        if "domain" not in cur or not cur["domain"]:
            cur["domain"] = _detect_domain(" ".join(lines), path.name)

        if "title" not in cur or not cur["title"]:
            cur["title"] = f"{path.stem} p{page_no}"

        # source_text 필수 필드 보강
        cur["source_text"] = cur.get("source_text") or " ".join(lines)[:4000]

        # payload_json dict 형태 고정 보장
        try:
            if isinstance(cur.get("payload_json"), str):
                _ = json.loads(cur["payload_json"])
            else:
                cur["payload_json"] = json.dumps(cur.get("payload_json", {}), ensure_ascii=False)
        except Exception:
            cur["payload_json"] = json.dumps({"raw": cur.get("payload_json")}, ensure_ascii=False)

        rows.append(cur)

    if not rows:
        rows.append(
            {
                "title": f"{path.name} (텍스트 추출 실패)",
                "domain": _detect_domain(path.stem, path.name),
                "tag": "pdf_import",
                "source_text": f"{path.name}에서 텍스트 추출 실패",
                "source_url": None,
                "payload_json": json.dumps({"source_file": path.name, "source_type": "pdf"}, ensure_ascii=False),
                "is_active": True,
            }
        )
    return rows


def read_source_file(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xls", ".excel"}:
        return read_xlsx(path)
    if suffix == ".pdf":
        return read_pdf(path)
    raise ValueError("지원 형식은 .csv, .xlsx, .pdf만 허용됩니다.")


async def upsert_rows(
    rows: Iterable[Dict[str, Any]],
    dry_run: bool,
    require_review_pass: bool,
    min_confidence: int,
    fail_on_quality_issue: bool,
) -> Dict[str, int]:
    summary = {
        "upserted": 0,
        "updated": 0,
        "inserted": 0,
        "failed": 0,
        "quality_warning": 0,
        "quality_rejected": 0,
    }

    async with AsyncSessionLocal() as session:
        # Ensure schema search path when run manually outside app startup config
        if settings.DATABASE_SCHEMA != "public":
            await session.execute(
                text(f'SET search_path TO "{settings.DATABASE_SCHEMA}", public')
            )

        for row in rows:
            normalized, errs, quality_warnings = _coerce_row(
                row,
                require_review_pass=require_review_pass,
                min_confidence=min_confidence,
            )
            if errs:
                summary["failed"] += 1
                print(f"[SKIP] {errs} row={row}")
                continue

            if quality_warnings:
                summary["quality_warning"] += 1
                print(f"[QUALITY_WARN] row={row.get('title')} {quality_warnings}")
                if fail_on_quality_issue and _is_blocked_quality_issue(quality_warnings):
                    summary["quality_rejected"] += 1
                    summary["failed"] += 1
                    continue

            stmt = select(ResearchStaticReferenceModel).where(
                and_(
                    ResearchStaticReferenceModel.domain == normalized["domain"],
                    ResearchStaticReferenceModel.industry_code == normalized["industry_code"],
                    ResearchStaticReferenceModel.tag == normalized["tag"],
                    ResearchStaticReferenceModel.title == normalized["title"],
                )
            )
            existing_row = (await session.execute(stmt)).scalar_one_or_none()

            if existing_row:
                for k, v in normalized.items():
                    setattr(existing_row, k, v)
                summary["updated"] += 1
            else:
                new_row = ResearchStaticReferenceModel(**normalized)
                session.add(new_row)
                summary["inserted"] += 1

            summary["upserted"] += 1

        if not dry_run:
            await session.commit()
        else:
            await session.rollback()

    summary["upserted"] = summary["inserted"] + summary["updated"]
    return summary


def main(argv: Optional[List[str]] = None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True, help="CSV/XLSX/PDF source file path")
    parser.add_argument("--dry-run", action="store_true", help="Dry run without DB write")
    parser.add_argument(
        "--require-review-pass",
        action="store_true",
        help="need_review=true인 행을 스킵 (CSV 템플릿 품질 게이트)",
    )
    parser.add_argument(
        "--min-confidence",
        type=int,
        default=0,
        help="최소 품질 confidence(0이면 비활성)",
    )
    parser.add_argument(
        "--fail-on-quality-issue",
        action="store_true",
        help="품질 경고 행을 치명 오류로 처리",
    )
    args = parser.parse_args(argv)

    source_path = Path(args.source).expanduser()
    if not source_path.exists():
        raise FileNotFoundError(f"source not found: {source_path}")

    rows = read_source_file(source_path)
    result = asyncio.run(
        upsert_rows(
            rows,
            dry_run=args.dry_run,
            require_review_pass=args.require_review_pass,
            min_confidence=args.min_confidence,
            fail_on_quality_issue=args.fail_on_quality_issue,
        )
    )
    print(f"[RESULT] total={len(rows)} {result}")


if __name__ == "__main__":
    main()
